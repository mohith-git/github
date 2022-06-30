# -*- coding: utf-8 -*-
"""
Created on Sun Jun 26 14:06:43 2022

@author: MOHITHTHAMANA
"""

import xml.etree.ElementTree as ET
import numpy as np









class TransMembrane:
    def __init__(self):
        self.xa=[]
        self.xb = []
        self.sr=[]
        self.mr=[]
        self.lr=[]
        self.lro=[]
        self.total_count={}
        self.cys1=[]
        self.aromatic = []
        self.aromatic_trp=[]
        self.arm=[]
        self.sulphur=[]
        self.cys=[]
        self.cation = []
        self.anion = []
        self.seq_xml=[]
        self.m_n = []
        self.m_o = []
        self.m_s = []
    def reading_file(self):
        f = open("E:/MOHITH/DDP/7a0x.pdb")
        a=f.read()
        a=a.splitlines()
        param = input("Chain Name1").upper()
        xa=[]
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2]=='CA':
                self.xa.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2]=='CB':
                self.xb.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2]=='SG':
                self.cys1.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2] in ['CG','CD1','CD2','CE1','CE2','CZ'] and aa[3] in ['PHE','TYR']:
                self.aromatic.append(aa)
            if aa[0]=='ATOM' and aa[4]==param and aa[2] in ['CD2','CZ2','CZ3','CE2','CE3','CH2'] and aa[3] in ['TRP']:
                self.aromatic_trp.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2] in ['SG','SD'] and aa[3] in ['MET','CYS']:
                self.sulphur.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break   
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2]=='SG' and aa[3]=='CYS':
                self.cys.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break  
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2] in ['NH1','NH2','NZ','ND1','NE2'] and aa[3] in ['ARG','LYS','HIS']:
                self.cation.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break  
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2] in ['OE2','OE1','OD1','OD2'] and aa[3] in ['ASP','GLU']:
                self.anion.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break 
        for i in a:
            aa=i.split()
            if aa[0]=='ATOM' and aa[4]==param and aa[2] in ['N','NE','NH1','NH2','NZ','ND1','NE2','NE1','ND2']:
                self.m_n.append(aa)
            elif aa[0]=='ATOM' and aa[4]==param and aa[2] in ['O','OE2','OE1','OD1','OD2','OG1','OG2','OH','OG']:
                self.m_o.append(aa)
            elif aa[0]=='ATOM' and aa[4]==param and aa[2]==['SG','SD']:
                self.m_s.append(aa)
            if aa[0] == 'MODEL' and aa[1] == '2':
                break
        
        return self.xa
    def distance_formula(self,i,j):
        i[6]=float(i[6])
        i[7]=float(i[7])
        i[8]=float(i[8])
        j[6]=float(j[6])
        j[7]=float(j[7])
        j[8]=float(j[8])
        d=(i[6]-j[6])**2+(i[7]-j[7])**2+(i[8]-j[8])**2
        return d
    def Short_range(self):
        i1=0
        for i in self.xa:
            j1=0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if 3.5<d**0.5<7 and [j[3],j[5],i[3],i[5],round(d**0.5,2)] not in self.sr and abs(i1-j1)<=2 :
                    self.sr.append([i[3],i[5],j[3],j[5],round(d**0.5,2)])
                j1=j1+1
            i1=i1+1
        return self.sr
    def Medium_range(self):
        i1=0
        for i in self.xa:
            j1=0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if 3.5<d**0.5<7 and [j[3],j[5],i[3],i[5],round(d**0.5,2)] not in self.mr and 3<=abs(i1-j1)<=4 :
                    self.mr.append([i[3],i[5],j[3],j[5],round(d**0.5,2)])
                j1=j1+1
            i1=i1+1
        return self.mr
    def Long_range(self):
        i1=0
        for i in self.xa:
            j1=0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if 3.5<d**0.5<7 and [j[3],j[5],i[3],i[5],round(d**0.5,2)] not in self.lr and abs(i1-j1)>4 :
                    self.lr.append([i[3],i[5],j[3],j[5],round(d**0.5,2)])
                j1=j1+1
            i1=i1+1
        return self.lr
    def LRO(self):
        i1=0
        for i in self.xa:
            j1=0
            s=0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if d**0.5<8 and abs(i1-j1)>12 :
                    s=s+(1/len(self.xa))
                j1=j1+1
            self.lro.append(s)
            i1=i1+1
        return self.lro
    def Contact_Order(self):
        i1=0
        c_order=0
        l=0
        summ=0
        for i in self.xa:
            j1=0
            c_order=0
            l=0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if d**0.5<8:
                    l=l+1
                c_order=c_order+(abs(j1-i1))
                j1=j1+1
            summ=summ+(c_order/l)
            i1=i1+1
        return(summ/len(self.xa))
    def Contact_Degree(self):
        degree = {}
    
        for i in self.xb:
            j1=0
            deg = 0
            for j in self.xb:
                d=c.distance_formula(i,j)
                if d**0.5<8:
                    deg+=1
                j1=j1+1
            degree[int(i[5])] = deg
        for i in self.xa:
          if i[3] == 'GLY': 
            j1=0
            deg = 0
            for j in self.xa:
               
                d=c.distance_formula(i,j)
                if 0<d**0.5<8:
                    deg+=1
                    j1=j1+1
            degree[int(i[5])] = deg
       
        for key in sorted(degree):
           return(key, degree[key])
    def A8(self):
        inp = float(input())
        degree = {}
        for i in self.xa:
            j1=0
            deg = 0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if 0<d**0.5<=inp:
                    deg+=1
                j1=j1+1
            degree['{0}-{1}'.format(i[3],i[5])] = deg
        #print('C-alpha:{0}'.format(inp))
        return(degree)
    # def A14(self):
    #     inp = float(input())
    #     degree = {}
    #     for i in self.xa:
    #         j1=0
    #         deg = 0
    #         for j in self.xa:
    #             d=c.distance_formula(i,j)
    #             if 0<d**0.5<=inp:
    #                 deg+=1
    #             j1=j1+1
    #         degree['{0}-{1}'.format(i[3],i[5])] = deg
    #     print(degree)
    def B8(self):
        inp = float(input())
        degree = {}
        for i in self.xb:
            j1=0
            deg = 0
            for j in self.xb:
                d=c.distance_formula(i,j)
                if 0<d**0.5<=inp:
                    deg+=1
                j1=j1+1
            degree['{0}-{1}'.format(i[3],i[5])] = deg
        #print('C-beta:{0}'.format(inp))
        return (degree)
    # def B14(self):
    #     inp = float(input())
    #     degree = {}
    #     for i in self.xb:
    #         j1=0
    #         deg = 0
    #         for j in self.xb:
    #             d=c.distance_formula(i,j)
    #             if 0<d**0.5<=inp:
    #                 deg+=1
    #             j1=j1+1
    #         degree['{0}-{1}'.format(i[3],i[5])] = deg
    #     print(degree)
    def disulphide_interaction(self):
        cys=[]
        for i in self.cys1:
            if i[3]=='CYS':
                cys.append(i)
        print(cys)
        p=[]
        i1=0
        for i in cys:
            j1=0
            for j in cys:
                d=c.distance_formula(i,j)
                if 0<d**0.5<=2.2 and [j[3],j[5],i[3],i[5],d**0.5] not in p:
                    p.append([i[3],i[5],j[3],j[5],d**0.5])
                j1=j1+1
            i1=i1+1
        return(p)
    def Hydrophobic_interaction(self):
        
        p=[]
        i1=0
        for i in self.xb:
          if i[3] in ['ALA','VAL','LEU','TYR','PRO','TRP','MET','PHE','ILE']:
            j1=0
            for j in self.xb:
              if j[3] in ['ALA','VAL','LEU','TYR','PRO','TRP','MET','PHE','ILE']:
                d=c.distance_formula(i,j)
                if 0<d**0.5<=5 and [j[3],j[5],i[3],i[5],d**0.5] not in p:
                    p.append([i[3],i[5],j[3],j[5],d**0.5])
                j1=j1+1
            i1=i1+1
        return(p)
        
    def Aromatic_Aromatic(self):
        p=[]
        for i in range(0,len(self.aromatic)-6,6):
            px=(float(self.aromatic[i][6])+float(self.aromatic[i+1][6])+float(self.aromatic[i+2][6])+float(self.aromatic[i+3][6])+float(self.aromatic[i+4][6])+float(self.aromatic[i+5][6]))/6
            py=(float(self.aromatic[i][7])+float(self.aromatic[i+1][7])+float(self.aromatic[i+2][7])+float(self.aromatic[i+3][7])+float(self.aromatic[i+4][7])+float(self.aromatic[i+5][7]))/6
            pz=(float(self.aromatic[i][8])+float(self.aromatic[i+1][8])+float(self.aromatic[i+2][8])+float(self.aromatic[i+3][8])+float(self.aromatic[i+4][8])+float(self.aromatic[i+5][8]))/6
            self.arm.append([self.aromatic[i][3],self.aromatic[i][5],px,py,pz])
        for i in range(0,len(self.aromatic_trp)-6,6):
            px=(float(self.aromatic_trp[i][6])+float(self.aromatic_trp[i+1][6])+float(self.aromatic_trp[i+2][6])+float(self.aromatic_trp[i+3][6])+float(self.aromatic_trp[i+4][6])+float(self.aromatic_trp[i+5][6]))/6
            py=(float(self.aromatic_trp[i][7])+float(self.aromatic_trp[i+1][7])+float(self.aromatic_trp[i+2][7])+float(self.aromatic_trp[i+3][7])+float(self.aromatic_trp[i+4][7])+float(self.aromatic_trp[i+5][7]))/6
            pz=(float(self.aromatic_trp[i][8])+float(self.aromatic_trp[i+1][8])+float(self.aromatic_trp[i+2][8])+float(self.aromatic_trp[i+3][8])+float(self.aromatic_trp[i+4][8])+float(self.aromatic_trp[i+5][8]))/6
            self.arm.append([self.aromatic_trp[i][3],self.aromatic_trp[i][5],px,py,pz])
        for i in self.arm:
            for j in self.arm:
                d=((i[2]-j[2])**2+(i[3]-j[3])**2+(i[4]-j[4])**2)**0.5
                if 4.5<d<7 and [j[0],j[1],i[0],i[1],d**0.5] not in p:
                    p.append([i[0],i[1],j[0],j[1],d**0.5])
        return p
    def Aromatic_sulphur(self):
        p=[]
        for i in self.arm:
            for j in self.sulphur:
                d=((i[2]-float(j[6]))**2+(i[3]-float(j[7]))**2+(i[4]-float(j[8]))**2)**0.5
                if d<5.3 and [j[3],j[5],i[0],i[1],d**0.5] not in p:
                    p.append([i[0],i[1],j[3],j[5],d**0.5])
        return p
    def disulphide(self):
        dis=[]
        i1=0
        for i in self.cys:
            j1=0
            for j in self.cys:
                d=c.distance_formula(i,j)
                if 0<d**0.5<2.2 and [j[3],j[5],i[3],i[5],d**0.5] not in dis:
                    dis.append([i[3],i[5],j[3],j[5],d**0.5])
                j1=j1+1
            i1=i1+1
        return(dis)
    def Ionic_Interaction(self):
        ion = []
        i1=0
        for i in self.cation:
            j1=0
           
            for j in self.anion:
                d=c.distance_formula(i,j)
                if 0<d**0.5<6:
                    ion.append([i[3],i[5],i[2],j[3],j[5],j[2],round(d**0.5,2)])
                j1=j1+1
            i1=i1+1
         
        
        
        return(ion)
           
    def Cation_pi(self):
        cpi=[]
        i1=0
        for i in self.cation:
         if i[3]!='HIS':
            j1=0
            for j in self.arm:
                d=(float(i[6])-j[2])**2+(float(i[7])-j[3])**2+(float(i[8])-j[4])**2
                if 0<d**0.5<10 and [j[0],j[1],i[3],i[5],i[2],d**0.5] not in cpi:
                    cpi.append([i[3],i[5],i[2],j[0],j[1],d**0.5])
                j1=j1+1
            i1=i1+1
        return((cpi))
    
    def Main_Main_H_Bond(self):
        h_bond = []
        i1=0
        for i in self.m_n:
         if i[2] == 'N':
            j1=0
            
            for j in self.m_o:
              if j[2] == 'O':
               if abs(int(i[5]) - int(j[5])) > 1:
                d=c.distance_formula(i,j)
                if 0<d**0.5<3.5 and [j[3],j[5],j[2],i[3],i[5],i[2],round(d**0.5,2)] not in h_bond:
                    h_bond.append([i[3],i[5],i[2],j[3],j[5],j[2],round(d**0.5,2)])
                j1=j1+1
            i1=i1+1
        h_bond = sorted(h_bond,key=lambda x : x[1])
        return((h_bond))
    
    
    def Main_Side_Chain_H_Bond(self):
        
        atoms= [*self.m_n,*self.m_o,*self.m_s]
        h_bond = []
        real = []
        i1=0
        for i in atoms:
           
            j1=0
            
            
            for j in atoms:
             
               if abs(int(i[5]) - int(j[5])) > 1:
                
                 d=c.distance_formula(i,j)
                
                 if 0<d**0.5<3.5:
                    if i[2] == 'O':
                      h_bond.append([j[3],j[5],j[2],i[3],i[5],i[2],round(d**0.5,2)])
                    else:
                        if j[2] == 'N':
                          h_bond.append([j[3],j[5],j[2],i[3],i[5],i[2],round(d**0.5,2)])   
                     
               j1=j1+1
            i1=i1+1
        for i in h_bond:
           if i[2] in ['N','O'] and i[5]in ['N','O']:
               pass
           elif i[2] in ['N','O'] or i[5]in ['N','O']:
               real.append(i)
        real = sorted(real,key=lambda x : x[1])
        return(((real)))   
    
    
    
    def Side_Side_Chain_H_Bond(self):
        atoms= [*self.m_n,*self.m_o,*self.m_s]
        h_bond = []
        
        i1=0
        for i in atoms:
           if i[2] not in ['N','O']:
            j1=0
            
            
            for j in atoms:
              if j[2] not in ['N','O']:
               if abs(int(i[5]) - int(j[5])) > 1:
                d=c.distance_formula(i,j)
                if 0<d**0.5<3.5 and [j[3],j[5],j[2],i[3],i[5],i[2],round(d**0.5,2)] not in h_bond:
                    h_bond.append([i[3],i[5],i[2],j[3],j[5],j[2],round(d**0.5,2)])
                j1=j1+1
            i1=i1+1
        
        return(((h_bond)))   
    
    
    
    def Surr_Hydrophob(self):
        hydro_index = {'ALA': 0.87 , 'ASP': 0.66, 'CYS': 1.52, 'GLU': 0.67, 'PHE': 2.87 , 'GLY':0.1, 'HIS': 0.87, 'ILE': 3.15,
       			'LYS': 1.64, 'LEU': 2.17,'MET':1.67 ,'ASN': 0.09,'PRO': 2.77,'GLN': 0 ,'ARG':0.85, 'SER': 0.07,
       			'THR':0.07, 'VAL':1.87, 'TRP': 3.77, 'TYR': 2.67}
        
        surr = []
        
        i1=0
        for i in self.xa:
            j1=0
            interact = []
            score = 0
            for j in self.xa:
                d=c.distance_formula(i,j)
                if 0<d**0.5<8:
                    interact.append(j[3])
                
            for k in set(interact):
                   if k=='AHIS' or k=='BHIS':
                    score+=(interact.count(k)*0.87)
                   else:
                    score+=(interact.count(k)*hydro_index[k]) 

            j1=j1+1
            surr.append([i[3],i[5],round(score,2)])
                
            i1=i1+1
         
        return((surr))
              
    def read_xml(self):
        import xml.etree.ElementTree as ET
        tree = ET.parse('E:/MOHITH/DDP/7a0x.xml')
        root = tree.getroot()
        for i in range(len(root)):
            if 'CHAIN' in root[i].tag and root[i].attrib['CHAINID']=='A':
                for j in range(len(root[i])):
                    if j!=0:
                        self.seq_xml.append(root[i][j].attrib)
        print(self.seq_xml)
    
    def Regions(self):
        extr,intr,helix= [],[],[]
        for i in self.seq_xml:
            if i['type'] == '1':
                for k in range(int(i['seq_beg']),int(i['seq_end'])+1):
                	extr.append(k)
            elif i['type'] == 'H': 
                for k in range(int(i['seq_beg']),int(i['seq_end'])+1):
                	helix.append(k)
            else:
                for k in range(int(i['seq_beg']),int(i['seq_end'])+1): 
                	intr.append(k)
        print('Extra Cellular:',extr)
        print('\n')
        print('alpha-helix:',helix)
        print('\n')
        print('Intra Cellular:',intr)
        return [extr,helix,intr]


    def excel_append(self,df,name):
        import pandas as pd
        import openpyxl
        dk=[]
        for i in df:
            print(i)
            dk.append(pd.DataFrame(i))
        with pd.ExcelWriter('pandas_to_excel.xlsx') as writer:
        	j=0
        	for i in dk:
        		i.to_excel(writer, sheet_name=name[j])
        		j=j+1
    def excel_coloring(self,xml,name):
        import openpyxl 
        from openpyxl.styles import PatternFill
        from xlrd import open_workbook
        wb = openpyxl.load_workbook("E:/MOHITH/DDP/pandas_to_excel.xlsx")
        spread_sheet=open_workbook("E:/MOHITH/DDP/pandas_to_excel.xlsx")
        sheet1=spread_sheet.sheet_by_index(0)
        for i in name:
        	ws = wb[i]
        	fill_cell1 = PatternFill(patternType='solid', 
                           fgColor='FC2C03')
        	fill_cell2 = PatternFill(patternType='solid', 
                           fgColor='03FCF4') 
        	fill_cell3 = PatternFill(patternType='solid', 
                           fgColor='35FC03')
        	j=0
        	for i in sheet1.col_values(2):
        		try:
        			if j>2:
        				if float(i) in xml[0]:
        					ws['C'+str(j)].fill = fill_cell1
        				elif float(i) in xml[1]:
        					ws['C'+str(j)].fill = fill_cell2
        				else:
        					ws['C'+str(j)].fill = fill_cell3
        		except:
        			pass
        		j=j+1
        	j=0
        	for i in sheet1.col_values(4):
        		try:
        			if j>2:
        				if float(i) in xml[0]:
        					ws['E'+str(j)].fill = fill_cell1
        				elif float(i) in xml[1]:
        					ws['E'+str(j)].fill = fill_cell2
        				else:
        					ws['E'+str(j)].fill = fill_cell3
        		except:
        			pass
        		j=j+1
        	wb.save("E:/MOHITH/DDP/pandas_to_excel.xlsx")
 	#elif aa[0]=='ATOM' and aa[4]==param[1]:
	     #xb.append(aa)
         
         
         
         




c=TransMembrane()
xa=c.reading_file()
sr=c.Short_range()
mr=c.Medium_range()
lr=c.Long_range()
mm=c.Main_Main_H_Bond()
ms=c.Main_Side_Chain_H_Bond()
ss=c.Side_Side_Chain_H_Bond()
lro=c.LRO()

co=c.Contact_Order()
#print(sr)
c.Contact_Degree()
c.A8()
di=c.disulphide_interaction()
hi=c.Hydrophobic_interaction()
ar=c.Aromatic_Aromatic()
ars=c.Aromatic_sulphur()
dis=c.disulphide()
ion=c.Ionic_Interaction()

cpi=c.Cation_pi()
# #sh=c.Surr_Hydrophob()
# c.read_xml()
# a=c.Regions()
# name=['short_range','medium_range','long_range',"LRO","disulphide_interaction_cystine","Hydrophobic_interaction","Aromatic_Aromatic","Aromatic_sulphur","disulphide_interaction","Ionic_Interaction","Cation_pi",'Main_Main_Hydrogen','Side_Main_Hydrogen','Side_Side_Chain']
# c.excel_append([sr,mr,lr,lro,di,hi,ar,ars,dis,ion,cpi,mm,ms,ss],name)
# c.excel_coloring(a,name)
# name=['short_range','medium_range','long_range',"LRO","disulphide_interaction_cystine","Hydrophobic_interaction","Aromatic_Aromatic","Aromatic_sulphur","disulphide_interaction","Ionic_Interaction","Cation_pi"]
# c.excel_append([sr,mr,lr,lro,di,hi,ar,ars,dis,ion,cpi],name)
# c.excel_coloring(a,name)














